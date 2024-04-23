from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from openpyxl import load_workbook, Workbook
from datetime import datetime
import time


email = REDACTED
password = REDACTED

def scrape_controlup():
    # get url, create webdriver, open and log into website
    url = REDACTED
    driver = webdriver.Firefox()
    driver.get(url)
    time.sleep(2)
    driver.find_element(By.CLASS_NAME, "dex_OptionButton_optionButton__d2a4c").click()
    time.sleep(2)
    driver.find_element(By.ID, "i0116").send_keys(email + "\n")
    driver.find_element(By.ID, "i0116").send_keys(Keys.ENTER)
    time.sleep(2)
    driver.find_element(By.ID, "i0118").send_keys(password + "\n")
    driver.find_element(By.ID, "i0118").send_keys(Keys.ENTER)
    time.sleep(2)
    driver.find_element(By.ID, "idSIButton9").click()
    time.sleep(30)

    # website works with iframes, switch to said frame to find the DOM elements
    iframe_element = driver.find_element(By.CSS_SELECTOR, "iframe[title^='edge'][title$='devices']")
    driver.switch_to.frame(iframe_element)
    
    user_list = []

    time.sleep(2)

    # find the "online" filter, click it to ensure only online devices show
    header_row = driver.find_element(By.CLASS_NAME, "styles_headerRow__lC3VS").find_elements(By.TAG_NAME, "th")[3]
    online_input = header_row.find_elements(By.TAG_NAME, "div")[6]
    online_input.click()
    driver.find_element(By.ID, "menu-").find_elements(By.TAG_NAME, "li")[1].click()

    time.sleep(2)

    # get number of pages, start loop to go through all pages
    # todo
    number_of_pages = driver.find_element(By.CLASS_NAME, "styles_pageNumberContainer__BCXtE").find_elements(By.TAG_NAME, "b")[0].text
    page_input = driver.find_element(By.ID, "table-page-index-input")
    for i in range(2,int(number_of_pages) + 2):

    # loop through each row in the table, grab the username.
    # remove the BLACKSUN\ part of the username, if it is there
    # append to user list variable
        time.sleep(1)
        
        device_list = driver.find_element(By.CLASS_NAME, "styles_body__XCq05")
        devices = device_list.find_elements(By.TAG_NAME, "tr")
        for device in devices:
            device_children = device.find_elements(By.TAG_NAME, "td")
            device_text = device_children[8].text

            # format the names, first remove blacksun\ if the name contains it
            if device_text[0:8] == "BLACKSUN" or device_text[0:7] == "AzureAD":
                user_list.append(device_text.split("\\")[1])
            else:
                user_list.append(device_text)

        page_input.clear()
        page_input.send_keys(i)
        
        time.sleep(1)

    driver.close()
    return user_list

def export_to_csv(file_path, users):
    current_datetime = datetime.now()
    print(users)
    print(current_datetime)
    
    try:
        wb = load_workbook(file_path)
        ws = wb["OFFICE_PRESENCE"]
        
        last_row = ws.max_row

        for index, value in enumerate(users, start=1):
            ws.cell(row=last_row + index, column=1, value=value)
            ws.cell(row=last_row + index, column=2, value=current_datetime)

        wb.save(file_path)
        wb.close()
    
    except FileNotFoundError:
        wb = Workbook()
        ws = wb.active
        ws.title = "OFFICE_PRESENCE"

        ws["A1"] = "Users"
        ws["B1"] = "DateTime"

        for index, value in enumerate(users, start=2):
            ws.cell(row=index, column=1, value=value)
            ws.cell(row=index, column=2, value=current_datetime)

        wb.save(file_path)
        wb.close()

if __name__ == "__main__":
    filename = REDACTED
    users = scrape_controlup()

    new_users = []
    for name in users:
        name_split = name.split(" ")

        if len(name_split) > 1:
            new_name = name_split[0][0] + name_split[-1]  # First initial + last name
            new_users.append(new_name.lower())
        else:
            new_users.append(name.lower())  # Keep the name unchanged if it doesn't contain a last name

    users = new_users

    export_to_csv(filename, users)
